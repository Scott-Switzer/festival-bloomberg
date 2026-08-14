"""Design Partner multi-file ingestion: CSV, TSV, XLSX.

Pipeline:

    RAW FILE
    -> column mapping (conservative; ambiguous -> REVIEW_REQUIRED)
    -> PII quarantine (prohibited buyer PII never read)
    -> deterministic canonical event key (customer_event_id, or artist+venue+date)
    -> outcome claims (OBSERVED_PRIVATE, idempotent, append-only)
    -> data-quality audit + accounting reconciliation

Private observations are never merged with public ones; pooling is opt-in via
the dataset sharing policy (default PRIVATE_ONLY).
"""

from __future__ import annotations

import csv
import re
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from ..acquisition.contracts import content_hash_of, utc_now
from .design_partner import (
    AUTO_ACCEPTED,
    PROHIBITED,
    POTENTIAL_PII,
    REVIEW_REQUIRED,
    SAFE,
    SHARING_PRIVATE_ONLY,
    SHARING_POLICIES,
    ColumnMapping,
    classify_pii,
    map_columns,
)
from .retrospective import DEFAULT_ALLOWED_PRIVATE_INPUTS
from .outcome_claims import (
    ARTIST_BACKEND,
    ARTIST_GUARANTEE,
    COMP_TICKETS,
    EVENT_USABLE_CAPACITY,
    EXPLICIT_NOT_SOLD_OUT_ASSERTION,
    EXPLICIT_SOLD_OUT_ASSERTION,
    FNB_REVENUE,
    GRADE_A_PRIMARY_PROMOTER,
    GRADE_A_PRIMARY_SETTLEMENT,
    GRADE_A_PRIMARY_TICKETING,
    GRADE_A_PRIMARY_VENUE,
    LABOR_COST,
    MARKETING_SPEND,
    MERCH_REVENUE,
    OBSERVED_PRIVATE,
    PAID_ATTENDANCE,
    PAID_TICKETS,
    PARKING_REVENUE,
    PRIMARY_FACE_VALUE_MAX,
    PRIMARY_FACE_VALUE_MIN,
    PRODUCTION_COST,
    PROMOTER_CONTRIBUTION,
    REFUNDED_TICKETS,
    REPORTED_ATTENDANCE,
    RIGHTS_UNKNOWN,
    SCANNED_ATTENDANCE,
    SETTLEMENT_GROSS,
    SETTLEMENT_NET,
    SPONSOR_REVENUE,
    TICKET_GROSS,
    TICKET_NET,
    TICKETS_SOLD,
    VENUE_CAPACITY,
    VENUE_COST,
    VIP_REVENUE,
    OutcomeClaim,
)

# canonical field -> (outcome_type, source_grade). Only fields with a clean,
# unambiguous taxonomy mapping are written as claims; everything else is
# audited but not fabricated into a claim.
OUTCOME_FIELD_MAP: dict[str, tuple[str, str]] = {
    "tickets_sold": (TICKETS_SOLD, GRADE_A_PRIMARY_TICKETING),
    "paid_tickets": (PAID_TICKETS, GRADE_A_PRIMARY_TICKETING),
    "comp_tickets": (COMP_TICKETS, GRADE_A_PRIMARY_TICKETING),
    "refunded_tickets": (REFUNDED_TICKETS, GRADE_A_PRIMARY_TICKETING),
    "scanned_attendance": (SCANNED_ATTENDANCE, GRADE_A_PRIMARY_TICKETING),
    "paid_attendance": (PAID_ATTENDANCE, GRADE_A_PRIMARY_TICKETING),
    "reported_attendance": (REPORTED_ATTENDANCE, GRADE_A_PRIMARY_TICKETING),
    "ticket_gross": (TICKET_GROSS, GRADE_A_PRIMARY_TICKETING),
    "ticket_net": (TICKET_NET, GRADE_A_PRIMARY_TICKETING),
    "face_value_min": (PRIMARY_FACE_VALUE_MIN, GRADE_A_PRIMARY_TICKETING),
    "face_value_max": (PRIMARY_FACE_VALUE_MAX, GRADE_A_PRIMARY_TICKETING),
    "venue_capacity": (VENUE_CAPACITY, GRADE_A_PRIMARY_VENUE),
    "event_usable_capacity": (EVENT_USABLE_CAPACITY, GRADE_A_PRIMARY_VENUE),
    "artist_guarantee": (ARTIST_GUARANTEE, GRADE_A_PRIMARY_PROMOTER),
    "artist_backend_pct": (ARTIST_BACKEND, GRADE_A_PRIMARY_PROMOTER),
    "marketing_spend": (MARKETING_SPEND, GRADE_A_PRIMARY_PROMOTER),
    "venue_cost": (VENUE_COST, GRADE_A_PRIMARY_PROMOTER),
    "production_cost": (PRODUCTION_COST, GRADE_A_PRIMARY_PROMOTER),
    "labor_cost": (LABOR_COST, GRADE_A_PRIMARY_PROMOTER),
    "merch_revenue": (MERCH_REVENUE, GRADE_A_PRIMARY_PROMOTER),
    "fnb_revenue": (FNB_REVENUE, GRADE_A_PRIMARY_PROMOTER),
    "parking_revenue": (PARKING_REVENUE, GRADE_A_PRIMARY_PROMOTER),
    "vip_revenue": (VIP_REVENUE, GRADE_A_PRIMARY_PROMOTER),
    "sponsor_revenue": (SPONSOR_REVENUE, GRADE_A_PRIMARY_PROMOTER),
    "promoter_contribution": (PROMOTER_CONTRIBUTION, GRADE_A_PRIMARY_SETTLEMENT),
    "settlement_gross": (SETTLEMENT_GROSS, GRADE_A_PRIMARY_SETTLEMENT),
    "settlement_net": (SETTLEMENT_NET, GRADE_A_PRIMARY_SETTLEMENT),
}

_MONEY_FIELDS = frozenset({
    "artist_guarantee", "artist_bonus", "artist_expenses", "ticket_gross", "ticket_net",
    "average_paid_ticket", "face_value_min", "face_value_max", "marketing_spend",
    "venue_cost", "production_cost", "labor_cost", "security_cost", "insurance_cost",
    "other_cost", "merch_revenue", "fnb_revenue", "parking_revenue", "vip_revenue",
    "sponsor_revenue", "other_revenue", "promoter_contribution", "settlement_gross",
    "settlement_net",
})
_COUNT_FIELDS = frozenset({
    "venue_capacity", "event_usable_capacity", "ticket_capacity", "tickets_sold",
    "paid_tickets", "comp_tickets", "refunded_tickets", "scanned_attendance",
    "paid_attendance", "reported_attendance",
})


@dataclass
class PartnerIngestionReport:
    dataset_id: str = ""
    customer_id: str = ""
    files_read: int = 0
    rows_read: int = 0
    claims_built: int = 0
    claims_inserted: int = 0
    duplicates_skipped: int = 0
    pii_quarantined: int = 0
    events_resolved: list[str] = field(default_factory=list)
    mappings: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    pii: dict[str, dict[str, str]] = field(default_factory=dict)
    quality_issues: list[dict[str, Any]] = field(default_factory=list)
    reconciliation: list[dict[str, Any]] = field(default_factory=list)
    errors: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "dataset_id": self.dataset_id,
            "customer_id": self.customer_id,
            "files_read": self.files_read,
            "rows_read": self.rows_read,
            "claims_built": self.claims_built,
            "claims_inserted": self.claims_inserted,
            "duplicates_skipped": self.duplicates_skipped,
            "pii_quarantined": self.pii_quarantined,
            "events_resolved": sorted(set(self.events_resolved)),
            "mappings": self.mappings,
            "pii": self.pii,
            "quality_issues": self.quality_issues,
            "reconciliation": self.reconciliation,
            "errors": self.errors,
        }


# ---------------------------------------------------------------------------
# Reading tabular files (CSV / TSV / XLSX)
# ---------------------------------------------------------------------------
def _detect_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        return dialect.delimiter
    except csv.Error:
        return ","


def _read_delimited(path: Path, delimiter: str) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        delim = delimiter or _detect_delimiter(sample)
        reader = csv.DictReader(fh, delimiter=delim)
        headers = [h.strip() for h in (reader.fieldnames or [])]
        rows = [dict(row) for row in reader]
    return headers, rows


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if first is None:
            return [], []
        headers = [str(h).strip() if h is not None else "" for h in first]
        rows: list[dict[str, str]] = []
        for raw in rows_iter:
            row: dict[str, str] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = raw[index] if index < len(raw) else None
                row[header] = "" if value is None else str(value).strip()
            rows.append(row)
        return headers, rows
    finally:
        wb.close()


def read_tabular(path: str | Path) -> tuple[list[str], list[dict[str, str]]]:
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(p)
    if suffix in (".tsv", ".tab"):
        return _read_delimited(p, "\t")
    return _read_delimited(p, ",")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "").replace("$", "")
    if text in ("", "-", "n/a", "NA", "null", "None"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _first_present(row: dict[str, str], fields: list[str], mapping_by_canonical: dict[str, str]) -> str | None:
    for field in fields:
        value = _cell_value(row, field, mapping_by_canonical)
        if value is not None and str(value).strip():
            return str(value).strip()
    return None


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-") or "unknown"


def _cell_value(row: dict[str, str], canonical_field: str, mapping_by_canonical: dict[str, str]) -> Any:
    source_header = mapping_by_canonical.get(canonical_field)
    if source_header is None:
        return None
    return row.get(source_header)


def resolve_event_key(row: dict[str, str], *, customer_id: str, mapping_by_canonical: dict[str, str]) -> str:
    """Deterministic private-event key. Prefers customer_event_id; falls back
    to artist+venue+date. Never guesses an identity merge with the public
    graph here (that is the caller's job)."""
    external = _cell_value(row, "customer_event_id", mapping_by_canonical)
    if external and str(external).strip():
        return f"private_{customer_id}_{_slug(str(external).strip())}"
    artist = _cell_value(row, "artist_name", mapping_by_canonical)
    venue = _cell_value(row, "venue_name", mapping_by_canonical)
    date = _cell_value(row, "event_date", mapping_by_canonical)
    return f"private_{customer_id}_{_slug(str(artist or 'unknown'))}_{_slug(str(venue or 'unknown'))}_{str(date or '').strip()}"


def build_claims_for_row(
    row: dict[str, str],
    *,
    canonical_event_id: str,
    customer_id: str,
    dataset_id: str,
    source_file_id: str,
    row_number: int,
    mapping_by_canonical: dict[str, str],
    currency: str | None,
) -> list[OutcomeClaim]:
    claims: list[OutcomeClaim] = []
    # PIT knowledge times: inputs are knowable at announcement (or earlier),
    # realized outcomes only at the event. Documented approximation.
    input_knowledge = _first_present(
        row, ["announcement_date", "onsale_date", "booking_date", "offer_date"], mapping_by_canonical,
    )
    outcome_knowledge = _first_present(row, ["event_date"], mapping_by_canonical)

    for canonical_field, (outcome_type, grade) in OUTCOME_FIELD_MAP.items():
        raw = _cell_value(row, canonical_field, mapping_by_canonical)
        if raw is None or str(raw).strip() == "":
            continue
        value = _to_float(raw)
        if value is None:
            continue
        claim_id = f"claim_{content_hash_of({
            'provider': 'partner_import',
            'customer': customer_id,
            'dataset': dataset_id,
            'event': canonical_event_id,
            'type': outcome_type,
            'value': value,
        })[:20]}"
        notes = f"private import {source_file_id} row {row_number}"
        is_input = outcome_type in DEFAULT_ALLOWED_PRIVATE_INPUTS
        knowledge_time = input_knowledge if is_input else outcome_knowledge
        claims.append(
            OutcomeClaim.build(
                claim_id=claim_id,
                canonical_event_id=canonical_event_id,
                outcome_type=outcome_type,
                value_numeric=value,
                currency=currency if canonical_field in _MONEY_FIELDS else None,
                source_provider="design_partner",
                source_name=dataset_id,
                source_document_id=source_file_id,
                source_quality=grade,
                observation_class=OBSERVED_PRIVATE,
                rights_status=RIGHTS_UNKNOWN,
                commercial_use_status=RIGHTS_UNKNOWN,
                knowledge_time=knowledge_time,
                notes=notes,
            )
        )
    # Explicit sold-out flag becomes an assertion claim, never inferred.
    sold_out_raw = _cell_value(row, "sold_out", mapping_by_canonical)
    if sold_out_raw is not None and str(sold_out_raw).strip() != "":
        text = str(sold_out_raw).strip().lower()
        if text in ("true", "yes", "y", "1", "sold_out", "sold out"):
            outcome_type = EXPLICIT_SOLD_OUT_ASSERTION
        elif text in ("false", "no", "n", "0", "not_sold_out", "not sold out"):
            outcome_type = EXPLICIT_NOT_SOLD_OUT_ASSERTION
        else:
            outcome_type = None
        if outcome_type is not None:
            claims.append(
                OutcomeClaim.build(
                    claim_id=f"claim_{content_hash_of({
                        'provider': 'partner_import', 'customer': customer_id,
                        'event': canonical_event_id, 'type': outcome_type,
                    })[:20]}",
                    canonical_event_id=canonical_event_id,
                    outcome_type=outcome_type,
                    value_text=text,
                    source_provider="design_partner",
                    source_name=dataset_id,
                    source_document_id=source_file_id,
                    source_quality=GRADE_A_PRIMARY_PROMOTER,
                    observation_class=OBSERVED_PRIVATE,
                    rights_status=RIGHTS_UNKNOWN,
                    commercial_use_status=RIGHTS_UNKNOWN,
                    notes=f"explicit sold-out flag {source_file_id} row {row_number}",
                )
            )
    return claims


# ---------------------------------------------------------------------------
# Data quality audit
# ---------------------------------------------------------------------------
def data_quality_audit(
    rows: list[dict[str, str]],
    *,
    mapping_by_canonical: dict[str, str],
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        def num(field: str) -> float | None:
            return _to_float(_cell_value(row, field, mapping_by_canonical))

        currency = _cell_value(row, "currency", mapping_by_canonical)
        gross = num("ticket_gross")
        net = num("ticket_net")
        sold = num("tickets_sold")
        if sold is None:
            sold = num("paid_tickets")
        capacity = num("venue_capacity")
        if capacity is None:
            capacity = num("event_usable_capacity")
        scans = num("scanned_attendance")
        refunds = num("refunded_tickets")
        paid = num("paid_tickets")

        if gross is not None and gross < 0:
            issues.append({"row": index, "check": "negative_gross", "reason": "ticket_gross < 0"})
        if net is not None and net < 0:
            issues.append({"row": index, "check": "negative_net", "reason": "ticket_net < 0"})
        if net is not None and gross is not None and net > gross:
            issues.append({"row": index, "check": "net_exceeds_gross", "reason": "ticket_net > ticket_gross"})
        if sold is not None and capacity is not None and sold > capacity:
            issues.append({"row": index, "check": "tickets_exceed_capacity", "reason": "tickets_sold > capacity"})
        if scans is not None and sold is not None and scans > sold:
            issues.append({"row": index, "check": "scans_exceed_tickets", "reason": "scanned_attendance > tickets_sold"})
        if refunds is not None and sold is not None and refunds > sold:
            issues.append({"row": index, "check": "refunds_exceed_sold", "reason": "refunded_tickets > tickets_sold"})
        if paid is not None and paid < 0:
            issues.append({"row": index, "check": "negative_paid_tickets", "reason": "paid_tickets < 0"})
        if currency and str(currency).strip().upper() not in ("USD", "CAD", "EUR", "GBP"):
            issues.append({"row": index, "check": "unknown_currency", "reason": f"currency {currency}"})

    # Dataset-level: mixed currencies across rows are a reconciliation hazard.
    distinct_currencies = {
        str(_cell_value(row, "currency", mapping_by_canonical)).strip().upper()
        for row in rows
        if _cell_value(row, "currency", mapping_by_canonical) and str(_cell_value(row, "currency", mapping_by_canonical)).strip()
    }
    if len(distinct_currencies) > 1:
        issues.append({"row": None, "check": "mixed_currency", "reason": f"multiple currencies: {sorted(distinct_currencies)}"})
    return issues


# ---------------------------------------------------------------------------
# Accounting reconciliation
# ---------------------------------------------------------------------------
RECON_TIES = "TIES"
RECON_NEAR_TIE = "NEAR_TIE"
RECON_DOES_NOT_TIE = "DOES_NOT_TIE"
RECON_INSUFFICIENT = "INSUFFICIENT_FIELDS"


def accounting_reconciliation(
    rows: list[dict[str, str]],
    *,
    mapping_by_canonical: dict[str, str],
    tolerance_abs: float = 1.0,
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        def num(field: str) -> float | None:
            return _to_float(_cell_value(row, field, mapping_by_canonical))

        net = num("ticket_net")
        merch = num("merch_revenue")
        fnb = num("fnb_revenue")
        sponsor = num("sponsor_revenue")
        other_rev = num("other_revenue")
        guarantee = num("artist_guarantee")
        marketing = num("marketing_spend")
        venue_cost = num("venue_cost")
        production = num("production_cost")
        labor = num("labor_cost")
        other_cost = num("other_cost")
        promoter = num("promoter_contribution")

        revenue_components = [v for v in (net, merch, fnb, sponsor, other_rev) if v is not None]
        cost_components = [v for v in (guarantee, marketing, venue_cost, production, labor, other_cost) if v is not None]

        if promoter is None:
            results.append({"row": index, "status": RECON_INSUFFICIENT, "reason": "promoter_contribution missing"})
            continue
        if not revenue_components or not cost_components:
            results.append({"row": index, "status": RECON_INSUFFICIENT, "reason": "insufficient revenue/cost fields"})
            continue

        implied = sum(revenue_components) - sum(cost_components)
        delta = abs(implied - promoter)
        if delta <= tolerance_abs:
            status = RECON_TIES
        elif delta <= tolerance_abs * 10:
            status = RECON_NEAR_TIE
        else:
            status = RECON_DOES_NOT_TIE
        results.append({
            "row": index,
            "status": status,
            "implied_promoter_contribution": round(implied, 2),
            "reported_promoter_contribution": promoter,
            "difference": round(delta, 2),
            "tolerance_abs": tolerance_abs,
        })
    return results


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------
def _register_private_event(
    *,
    events_repo,
    row: dict[str, str],
    canonical_event_id: str,
    mapping_by_canonical: dict[str, str],
) -> None:
    """Register a private historical show in the canonical event master.

    Best-effort: if no event with this venue+date already exists, insert a
    minimal private-origin row (event_id = private key). The original customer
    identity is preserved in the source-file lineage, never here as a fake
    provider identity."""
    artist = _first_present(row, ["artist_name"], mapping_by_canonical) or "unknown"
    venue_name = _first_present(row, ["venue_name"], mapping_by_canonical)
    event_date = _first_present(row, ["event_date"], mapping_by_canonical)
    city = _first_present(row, ["city"], mapping_by_canonical)
    market = _first_present(row, ["market"], mapping_by_canonical)
    venue_id = f"ven_{_slug(str(venue_name or 'unknown'))}"

    existing = events_repo.conn.execute(
        "SELECT event_id FROM events.events WHERE event_id = ?", [canonical_event_id]
    ).fetchone()
    if existing:
        return
    now = utc_now().isoformat()
    events_repo.conn.execute(
        """
        INSERT INTO events.events
            (event_id, event_type, event_name, event_time, local_date, venue_id, venue_name,
             market_id, city, state, country, event_status, provider_support_count,
             first_observed_at, last_observed_at, knowledge_time, match_gate,
             supporting_observation_ids)
        VALUES (?, 'CONCERT', ?, NULL, ?, ?, ?, ?, ?, NULL, NULL, NULL, 1, ?, ?, ?, 'PRIVATE_IMPORT', ?)
        """,
        [
            canonical_event_id,
            artist,
            event_date,
            venue_id,
            venue_name,
            market,
            city,
            now,
            now,
            now,
            "[]",
        ],
    )
    events_repo.conn.execute(
        """
        INSERT INTO events.artist_event_relations
            (relation_id, artist_id, event_id, role, knowledge_time, supporting_observation_ids)
        VALUES (?, ?, ?, 'headliner', ?, ?)
        """,
        [f"aer_{canonical_event_id}", _slug(artist), canonical_event_id, now, "[]"],
    )
    events_repo.conn.commit()


def _write_decision_cutoffs(
    *,
    economics_repo,
    row: dict[str, str],
    canonical_event_id: str,
    mapping_by_canonical: dict[str, str],
) -> None:
    """Record per-event decision cutoffs from the imported decision dates."""
    booking = _first_present(row, ["booking_date", "offer_date"], mapping_by_canonical)
    announcement = _first_present(row, ["announcement_date"], mapping_by_canonical)
    onsale = _first_present(row, ["onsale_date", "presale_date"], mapping_by_canonical)
    event = _first_present(row, ["event_date"], mapping_by_canonical)
    economics_repo.upsert_decision_cutoffs({
        "event_id": canonical_event_id,
        "canonical_event_id": canonical_event_id,
        "booking_cutoff": booking,
        "announcement_cutoff": announcement,
        "onsale_cutoff": onsale,
        "event_cutoff": event,
        "cutoff_notes": "imported from design-partner file",
        "software_version": "design_partner_retrospective_v1",
    })


def ingest_partner_files(
    *,
    economics_repo,
    file_paths: list[str | Path],
    customer_id: str,
    dataset_id: str,
    sharing_policy: str = SHARING_PRIVATE_ONLY,
    source_system: str = "design_partner",
    link_to_public_events: bool = True,
    events_repo=None,
) -> PartnerIngestionReport:
    if sharing_policy not in SHARING_POLICIES:
        raise ValueError(f"invalid sharing policy {sharing_policy!r}")

    report = PartnerIngestionReport(dataset_id=dataset_id, customer_id=customer_id)
    economics_repo.create_customer_dataset(
        dataset_id=dataset_id,
        customer_id=customer_id,
        sharing_policy=sharing_policy,
        source_system=source_system,
    )
    ingestion_run_id = f"ir_{uuid.uuid4().hex[:16]}"
    economics_repo.insert_ingestion_run(
        ingestion_run_id=ingestion_run_id,
        dataset_id=dataset_id,
        software_version="design_partner_retrospective_v1",
    )

    for path in file_paths:
        p = Path(path)
        headers, rows = read_tabular(p)
        report.files_read += 1
        report.rows_read += len(rows)
        file_id = f"f_{content_hash_of(p.read_bytes())[:16]}"

        # Record source-file lineage (raw content hashed, not stored).
        economics_repo.insert_source_file(
            file_id=file_id,
            dataset_id=dataset_id,
            file_name=p.name,
            format=p.suffix.lstrip(".").lower(),
            row_count=len(rows),
            raw_content_hash=content_hash_of(p.read_bytes()),
        )

        # Column mapping + PII scan.
        mappings = map_columns(headers)
        report.mappings[p.name] = [m.to_dict() for m in mappings]
        pii = {h: classify_pii(h) for h in headers}
        report.pii[p.name] = pii

        mapping_by_canonical: dict[str, str] = {}
        for m in mappings:
            if m.status == AUTO_ACCEPTED and m.canonical_field and classify_pii(m.header) == SAFE:
                mapping_by_canonical[m.canonical_field] = m.header

        # Quarantine PII columns (values are never read into analytics).
        for header, status in pii.items():
            if status in (PROHIBITED, POTENTIAL_PII):
                economics_repo.insert_pii_quarantine(
                    quarantine_id=f"pq_{content_hash_of({'file': file_id, 'column': header})[:16]}",
                    file_id=file_id,
                    column_name=header,
                    reason="prohibited_buyer_pii" if status == PROHIBITED else "potential_pii_review_required",
                    sample_count=len(rows),
                )
                report.pii_quarantined += 1

        # Quality + reconciliation audits.
        report.quality_issues.extend(data_quality_audit(rows, mapping_by_canonical=mapping_by_canonical))
        report.reconciliation.extend(accounting_reconciliation(rows, mapping_by_canonical=mapping_by_canonical))

        # Build + write claims (idempotent, append-only).
        for index, row in enumerate(rows, start=1):
            canonical_event_id = resolve_event_key(
                row, customer_id=customer_id, mapping_by_canonical=mapping_by_canonical,
            )
            report.events_resolved.append(canonical_event_id)
            currency_raw = _cell_value(row, "currency", mapping_by_canonical)
            currency = str(currency_raw).strip().upper() if currency_raw else None

            for claim in build_claims_for_row(
                row,
                canonical_event_id=canonical_event_id,
                customer_id=customer_id,
                dataset_id=dataset_id,
                source_file_id=file_id,
                row_number=index,
                mapping_by_canonical=mapping_by_canonical,
                currency=currency,
            ):
                report.claims_built += 1
                if economics_repo.insert_outcome_claim(claim):
                    report.claims_inserted += 1
                else:
                    report.duplicates_skipped += 1

            _write_decision_cutoffs(
                economics_repo=economics_repo,
                row=row,
                canonical_event_id=canonical_event_id,
                mapping_by_canonical=mapping_by_canonical,
            )

            if events_repo is not None:
                _register_private_event(
                    events_repo=events_repo,
                    row=row,
                    canonical_event_id=canonical_event_id,
                    mapping_by_canonical=mapping_by_canonical,
                )

    return report
